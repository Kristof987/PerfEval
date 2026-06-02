from __future__ import annotations

import io
from copy import copy
from datetime import date
from typing import List, Dict, Any

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def auto_fit_columns(ws, min_width: int = 10, max_width: int = 80, padding: int = 2) -> None:
    for col_cells in ws.columns:
        max_length = 0
        col_index = col_cells[0].column
        col_letter = get_column_letter(col_index)
        for cell in col_cells:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        adjusted_width = min(max(max_length + padding, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width


def _eval_overall_5(ev: dict) -> float | None:
    vals = []
    for sec in ev.get("sections", []) or []:
        for q in sec.get("questions", []) or []:
            if not isinstance(q, dict) or q.get("type") != "rating":
                continue
            qid = str(q.get("id", ""))
            ans = (ev.get("answers", {}) or {}).get(qid)
            if isinstance(ans, dict):
                ans = ans.get("rating")
            if ans is None or ans == "":
                continue
            try:
                v = float(ans)
                m = float(q.get("rating_max", 5))
                if m > 0:
                    vals.append((v / m) * 5)
            except (ValueError, TypeError):
                pass
    if not vals:
        return None
    return sum(vals) / len(vals)


def _set_value_safe(sheet, coord: str, value) -> None:
    for rng in sheet.merged_cells.ranges:
        if coord in rng:
            sheet.cell(row=rng.min_row, column=rng.min_col).value = value
            return
    sheet[coord] = value


def _get_style_target_cell(sheet, coord: str):
    for rng in sheet.merged_cells.ranges:
        if coord in rng:
            return sheet.cell(row=rng.min_row, column=rng.min_col)
    return sheet[coord]


def _copy_row_style(sheet, src_row: int, dst_row: int) -> None:
    sheet.row_dimensions[dst_row].height = sheet.row_dimensions[src_row].height
    for col in range(1, sheet.max_column + 1):
        src_cell = sheet.cell(row=src_row, column=col)
        dst_cell = sheet.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)


def _ranges_intersect(a1: int, a2: int, a3: int, a4: int,
                      b1: int, b2: int, b3: int, b4: int) -> bool:
    return not (a3 < b1 or b3 < a1 or a4 < b2 or b4 < a2)


def _merge_block_safe(ws, block_start_row: int, block_end_row: int,
                      min_col: int, max_col: int) -> None:
    for rng in list(ws.merged_cells.ranges):
        if _ranges_intersect(
            min_col, block_start_row, max_col, block_end_row,
            rng.min_col, rng.min_row, rng.max_col, rng.max_row,
        ):
            ws.unmerge_cells(str(rng))
    ws.merge_cells(
        start_row=block_start_row,
        start_column=min_col,
        end_row=block_end_row,
        end_column=max_col,
    )


def _normalize_items(raw_list, key: str) -> List[str]:
    out = []
    seen: set = set()
    for it in raw_list or []:
        if not isinstance(it, dict):
            continue
        v = it.get(key, "")
        if isinstance(v, list):
            v = ", ".join(str(x).strip() for x in v if str(x).strip())
        txt = str(v).strip()
        if txt and txt.lower() not in seen:
            seen.add(txt.lower())
            out.append(txt)
    return out


def _set_summary_cell(ws, coord: str, title: str, items: List[str]) -> None:
    cell = _get_style_target_cell(ws, coord)
    bullet_text = "\n".join(f"• {x}" for x in (items or ["N/A"]))
    try:
        rich = CellRichText()
        rich.append(TextBlock(InlineFont(b=True), f"{title}\n"))
        rich.append(bullet_text)
        cell.value = rich
    except Exception:
        cell.value = f"{title}\n{bullet_text}"
    cell.alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
        shrink_to_fit=False,
    )


def build_ai_excel(
    emp_name: str,
    emp_email: str,
    emp_role: str,
    camp_name: str,
    submitted_count: int,
    evaluations: List[Dict[str, Any]],
    analysis: dict,
    template_path: str,
) -> bytes:
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Summary"] if "Summary" in wb.sheetnames else wb.active

    person_vals: Dict[str, list] = {}
    person_role: Dict[str, str] = {}
    for idx, ev in enumerate(evaluations):
        ov = _eval_overall_5(ev)
        if ov is None:
            continue
        name = (ev.get("evaluator_name") or "").strip() or f"evaluator_{idx}"
        role = (ev.get("evaluator_role") or "Unknown").strip() or "Unknown"
        person_vals.setdefault(name, []).append(ov)
        person_role.setdefault(name, role)

    person_avg: Dict[str, float] = {
        n: (sum(vs) / len(vs)) for n, vs in person_vals.items() if vs
    }
    manager_avgs = [
        avg for n, avg in person_avg.items()
        if "manager" in (person_role.get(n, "").lower())
    ]
    non_manager_avgs = [
        avg for n, avg in person_avg.items()
        if "manager" not in (person_role.get(n, "").lower())
    ]
    mgr_avg = round(sum(manager_avgs) / len(manager_avgs), 2) if manager_avgs else 0.0
    non_mgr_avg = round(sum(non_manager_avgs) / len(non_manager_avgs), 2) if non_manager_avgs else 0.0
    all_eval_avg = round(sum(person_avg.values()) / len(person_avg), 2) if person_avg else 0.0

    _set_value_safe(ws, "B5", emp_name)
    _set_value_safe(ws, "B6", emp_role if emp_role else "")
    _set_value_safe(ws, "B7", emp_email if emp_email else "")
    _set_value_safe(ws, "B11", date.today().strftime("%Y-%m-%d"))
    _set_value_safe(ws, "B12", camp_name if camp_name else "")
    _set_value_safe(ws, "B13", str(submitted_count))
    _set_value_safe(ws, "E12", mgr_avg)
    _set_value_safe(ws, "H7", non_mgr_avg)
    _set_value_safe(ws, "H12", round(non_mgr_avg - mgr_avg, 2))
    _set_value_safe(ws, "K7", all_eval_avg)

    section_ratings: Dict[str, list] = {}
    for ev in evaluations:
        for sec in ev.get("sections", []) or []:
            sec_title = (sec.get("title") or "General").strip() or "General"
            for q in sec.get("questions", []) or []:
                if not isinstance(q, dict) or q.get("type") != "rating":
                    continue
                qid = str(q.get("id", ""))
                ans = (ev.get("answers", {}) or {}).get(qid)
                if isinstance(ans, dict):
                    ans = ans.get("rating")
                if ans is None or ans == "":
                    continue
                try:
                    v = float(ans)
                    m = float(q.get("rating_max", 5))
                    if m > 0:
                        section_ratings.setdefault(sec_title, []).append((v, m))
                except (ValueError, TypeError):
                    pass

    section_names = list(section_ratings.keys())
    section_avgs_5 = [
        round(sum(v / m * 5 for v, m in rts) / len(rts), 2)
        for rts in section_ratings.values() if rts
    ]
    competencies = [str(c).strip() for c in section_names if str(c).strip()]

    if competencies:
        start_row = 17
        avg_by_comp = {
            str(name).strip(): avg
            for name, avg in zip(section_names, section_avgs_5)
        }
        header_row = start_row - 1
        avg_col = None
        for col_idx in range(1, ws.max_column + 1):
            hv = ws.cell(row=header_row, column=col_idx).value
            if isinstance(hv, str) and "average" in hv.strip().lower():
                avg_col = col_idx
                break
        if avg_col is None:
            avg_col = 2

        extra_rows = max(len(competencies) - 1, 0)
        if extra_rows > 0:
            saved_merges = []
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row >= start_row:
                    saved_merges.append((rng.min_col, rng.min_row, rng.max_col, rng.max_row))
                    ws.unmerge_cells(str(rng))
            ws.insert_rows(start_row + 1, amount=extra_rows)
            for mc, mr, xc, xr in saved_merges:
                new_min = mr + extra_rows if mr > start_row else mr
                new_max = xr + extra_rows if xr > start_row else xr
                if mr == start_row:
                    new_max = xr + extra_rows
                try:
                    ws.merge_cells(
                        start_row=new_min, start_column=mc,
                        end_row=new_max, end_column=xc,
                    )
                except Exception:
                    pass

        for idx in range(1, len(competencies)):
            _copy_row_style(ws, start_row, start_row + idx)

        for idx, comp_name in enumerate(competencies):
            row_idx = start_row + idx
            _set_value_safe(ws, f"A{row_idx}", comp_name)
            a_cell = _get_style_target_cell(ws, f"A{row_idx}")
            a_cell.alignment = Alignment(
                horizontal=a_cell.alignment.horizontal if a_cell.alignment else None,
                vertical=a_cell.alignment.vertical if a_cell.alignment else "center",
                text_rotation=a_cell.alignment.text_rotation if a_cell.alignment else 0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=a_cell.alignment.indent if a_cell.alignment else 0,
            )
            avg_val = avg_by_comp.get(comp_name)
            if avg_val is not None:
                avg_coord = f"{get_column_letter(avg_col)}{row_idx}"
                _set_value_safe(ws, avg_coord, float(avg_val))

        last_comp_row = start_row + len(competencies) - 1
        scan_from = last_comp_row + 1
        empty_count = 0
        for r in range(scan_from, scan_from + 20):
            cell_val = ws.cell(row=r, column=1).value
            if cell_val is None or str(cell_val).strip() == "":
                empty_count += 1
            else:
                break
        if empty_count > 0:
            ws.delete_rows(scan_from, amount=empty_count)

        block_start_row = start_row + len(competencies)
        block_end_row = block_start_row + 3
        _merge_block_safe(ws, block_start_row, block_end_row, 1, 5)
        _merge_block_safe(ws, block_start_row, block_end_row, 7, 11)

        top_strengths = analysis.get("top_strengths", [])
        top_dev = analysis.get("top_development_areas", [])
        strength_items = (
            [str(x).strip() for x in top_strengths if str(x).strip()]
            if isinstance(top_strengths, list) else []
        )
        dev_items = (
            [str(x).strip() for x in top_dev if str(x).strip()]
            if isinstance(top_dev, list) else []
        )
        _set_summary_cell(ws, f"A{block_start_row}", "Key Strengths", strength_items)
        _set_summary_cell(ws, f"G{block_start_row}", "Areas for Development", dev_items)

    fit_alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
        shrink_to_fit=True,
    )
    for cell_ref in ("B5", "B6", "B7", "B12"):
        _get_style_target_cell(ws, cell_ref).alignment = fit_alignment
    auto_fit_columns(ws)

    if "Részletes értékelések" in wb.sheetnames:
        wb.remove(wb["Részletes értékelések"])
    detail_sheet_name = "Detailed Answers"
    if detail_sheet_name in wb.sheetnames:
        ws_detail = wb[detail_sheet_name]
    else:
        ws_detail = wb.create_sheet(detail_sheet_name)

    ws_detail["A1"] = "Competence"
    ws_detail["B1"] = "Question"
    ws_detail["C1"] = "Question Type"
    ws_detail["D1"] = "Evaluator Role"
    ws_detail["E1"] = "Answer (text/non-rating)"
    ws_detail["F1"] = "Score (rating)"

    max_existing = max(ws_detail.max_row, 2)
    for row_idx in range(2, max_existing + 1):
        for col_idx in range(1, 7):
            ws_detail.cell(row=row_idx, column=col_idx).value = None

    row_out = 2
    for ev in evaluations:
        role = ev.get("evaluator_role", "")
        answers = ev.get("answers", {}) or {}
        for section in ev.get("sections", []) or []:
            category = section.get("title", "General")
            for q in section.get("questions", []) or []:
                if not isinstance(q, dict):
                    continue
                q_id = str(q.get("id", ""))
                q_text = q.get("text", "")
                q_type = q.get("type", "text")
                ans = answers.get(q_id)
                if isinstance(ans, dict):
                    ans = ans.get("rating") or ans.get("choice") or ans.get("text")
                ws_detail.cell(row=row_out, column=1).value = category
                ws_detail.cell(row=row_out, column=2).value = q_text
                ws_detail.cell(row=row_out, column=3).value = q_type
                ws_detail.cell(row=row_out, column=4).value = role
                if q_type in ("rating", "slider_labels"):
                    ws_detail.cell(row=row_out, column=5).value = ""
                    ws_detail.cell(row=row_out, column=6).value = "" if ans is None else str(ans)
                else:
                    ws_detail.cell(row=row_out, column=5).value = "" if ans is None else str(ans)
                    ws_detail.cell(row=row_out, column=6).value = ""
                row_out += 1

    auto_fit_columns(ws_detail)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
