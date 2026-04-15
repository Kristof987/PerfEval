import io
import os
import sys
import json
import subprocess
from copy import copy
from datetime import date, datetime
import openpyxl
import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from openpyxl.styles import Alignment
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter

from persistence.db.connection import get_db
from persistence.repository.campaign_repo import CampaignRepository

# Initialize repositories
campaign_repo = CampaignRepository()
db = get_db()

# -------------------------
# Session State Init
# -------------------------
if "cr_view" not in st.session_state:
    st.session_state.cr_view = "campaign"  # "campaign" or "employee" or "overall"
if "cr_selected_campaign_id" not in st.session_state:
    st.session_state.cr_selected_campaign_id = None
if "cr_selected_campaign_name" not in st.session_state:
    st.session_state.cr_selected_campaign_name = None
if "cr_selected_employee_id" not in st.session_state:
    st.session_state.cr_selected_employee_id = None
if "cr_selected_employee_name" not in st.session_state:
    st.session_state.cr_selected_employee_name = None


# -----------------------------------------------------------------------
# Helper: normalize f.questions → {"sections": [...]}
# -----------------------------------------------------------------------
def _normalize_questions(raw) -> dict:
    """
    f.questions can arrive as:
      - Python dict  (psycopg2 decoded jsonb)
      - Python list  (legacy flat question list)
      - JSON string  (text / json column)
      - None
    Always returns {"sections": [{"id": ..., "title": ..., "questions": [...]}]}
    """
    if raw is None:
        return {"sections": []}
    if isinstance(raw, str):
        raw = raw.strip()
        if not raw:
            return {"sections": []}
        try:
            raw = json.loads(raw)
        except (json.JSONDecodeError, ValueError):
            return {"sections": []}
    # legacy: flat list of question dicts
    if isinstance(raw, list):
        return {"sections": [{"id": "legacy", "title": "General", "questions": raw}]}
    # current: {"sections": [...]}
    if isinstance(raw, dict) and "sections" in raw:
        return raw
    return {"sections": []}


# -----------------------------------------------------------------------
# Helper: collect all completed evaluations for a campaign as a JSON-
# serialisable dict, ready to be passed to result_generation/main.py
# -----------------------------------------------------------------------
def get_campaign_qa_json(conn, campaign_id: int, campaign_name: str) -> dict:
    """
    Returns a dict structured as::

        {
            "campaign_id":   int,
            "campaign_name": str,
            "forms": {
                "<form_name>": [
                    {
                        "question":      str,
                        "question_type": str,
                        "competence":    str,
                        "options":       list,   # only for multiple_choice / slider_labels
                        "answers": [
                            {
                                "evaluatee_name": str,
                                "evaluator_role": str,
                                "answer":         any
                            },
                            ...
                        ]
                    },
                    ...
                ],
                ...
            }
        }
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT
            eval_tee.name          AS evaluatee_name,
            eval_tor.org_role_name AS evaluator_role,
            f.name                 AS form_name,
            e.answers,
            f.questions
        FROM evaluation e
        JOIN organisation_employees eval_tee ON e.evaluatee_id = eval_tee.id
        JOIN organisation_employees eval_tor ON e.evaluator_id = eval_tor.id
        JOIN form f ON e.form_id = f.id
        WHERE e.campaign_id = %s
          AND e.status = 'completed'
        ORDER BY f.name, eval_tee.name
    """, (campaign_id,))

    # form_name -> OrderedDict of q_id -> question entry (with "answers" list)
    form_questions: dict = {}

    for row in cur.fetchall():
        evaluatee_name = row[0]
        evaluator_role = row[1] or "Unknown"
        form_name      = row[2]
        answers        = row[3]
        questions_raw  = row[4]

        # --- parse answers ---
        if isinstance(answers, str):
            try:
                answers = json.loads(answers)
            except (json.JSONDecodeError, ValueError):
                answers = {}
        elif answers is None:
            answers = {}

        # --- parse questions into sections ---
        content  = _normalize_questions(questions_raw)
        sections = content.get("sections", [])

        # Build the question skeleton for this form on first encounter
        if form_name not in form_questions:
            form_questions[form_name] = {}
            for section in sections:
                section_title = section.get("title", "General")
                for q in section.get("questions", []):
                    if not isinstance(q, dict):
                        continue
                    q_id   = str(q.get("id", ""))
                    q_type = q.get("type", "text")

                    entry: dict = {
                        "question":      q.get("text", ""),
                        "question_type": q_type,
                        "competence":    section_title,
                        "answers":       [],
                    }
                    # Attach selectable options where relevant
                    if q_type == "multiple_choice":
                        entry["options"] = q.get("options", [])
                    elif q_type == "slider_labels":
                        entry["options"] = q.get("slider_options", [])

                    form_questions[form_name][q_id] = entry

        # --- attach each evaluator's answer to the matching question ---
        for q_id, q_entry in form_questions[form_name].items():
            answer_raw = answers.get(q_id)

            # Flatten legacy dict-wrapper answers
            if isinstance(answer_raw, dict):
                if "rating" in answer_raw:
                    answer_raw = answer_raw["rating"]
                elif "choice" in answer_raw:
                    answer_raw = answer_raw["choice"]
                elif "text" in answer_raw:
                    answer_raw = answer_raw["text"]

            if answer_raw is not None and answer_raw != "":
                q_entry["answers"].append({
                    "evaluatee_name": evaluatee_name,
                    "evaluator_role": evaluator_role,
                    "answer":         answer_raw,
                })

    cur.close()

    return {
        "campaign_id":   campaign_id,
        "campaign_name": campaign_name,
        "forms": {
            form_name: list(questions_by_id.values())
            for form_name, questions_by_id in form_questions.items()
        },
    }


# -----------------------------------------------------------------------
# Helper: render a single answer value given the question definition
# -----------------------------------------------------------------------
def _render_answer(answer, question: dict):
    q_type = question.get("type", "text")
    if answer is None or answer == "":
        st.caption("_(no answer)_")
        return
    # answers may be stored as a legacy dict wrapper {rating:…, choice:…, text:…}
    if isinstance(answer, dict):
        if "rating" in answer:
            hi = question.get("rating_max", 5)
            st.write(f"⭐ {answer['rating']}/{hi}")
        if "choice" in answer:
            st.write(f"☑️ {answer['choice']}")
        if "text" in answer:
            st.write(f"📝 {answer['text']}")
        return
    # scalar values (current storage format)
    if q_type == "rating":
        hi = question.get("rating_max", 5)
        st.write(f"⭐ {answer}/{hi}")
    elif q_type == "multiple_choice":
        st.write(f"☑️ {answer}")
    elif q_type == "slider_labels":
        st.write(f"🎚️ {answer}")
    else:
        st.write(f"📝 {answer}")


def _go_back_to_campaign_results():
    st.session_state.cr_view = "campaign"
    st.session_state.cr_selected_employee_id = None
    st.session_state.cr_selected_employee_name = None
    st.rerun()


def _render_cr_breadcrumb() -> None:
    parts = ["Campaign Results"]
    if st.session_state.cr_selected_campaign_name:
        parts.append(st.session_state.cr_selected_campaign_name)
    if st.session_state.cr_view == "overall":
        parts.append("Overall")
    elif st.session_state.cr_view == "employee" and st.session_state.cr_selected_employee_name:
        parts.append(st.session_state.cr_selected_employee_name)

    st.caption(" / ".join(parts))


def _participants_df_for_campaign(conn, campaign_id: int) -> pd.DataFrame:
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            oe.id,
            oe.name,
            oe.email,
            COALESCE(oe.org_role_name, 'No role') AS role,
            COALESCE(ev.completed_count, 0) AS completed_evaluations
        FROM organisation_employees oe
        JOIN employee_groups eg ON oe.id = eg.employee_id
        JOIN campaign_groups cg ON eg.group_id = cg.group_id
        LEFT JOIN (
            SELECT evaluatee_id, COUNT(*) AS completed_count
            FROM evaluation
            WHERE campaign_id = %s AND status = 'completed'
            GROUP BY evaluatee_id
        ) ev ON ev.evaluatee_id = oe.id
        WHERE cg.campaign_id = %s
        GROUP BY oe.id, oe.name, oe.email, oe.org_role_name, ev.completed_count
        ORDER BY oe.name ASC
        """,
        (campaign_id, campaign_id),
    )
    rows = cur.fetchall()
    cur.close()
    return pd.DataFrame(rows, columns=["id", "name", "email", "role", "completed_evaluations"])


def _render_summary_dashboard(evaluations: list, key_prefix: str = "overall"):
    if not evaluations:
        st.info("No completed evaluations found for this campaign.")
        return

    if len(evaluations) < 3:
        st.warning("Low data quality: fewer than 3 completed evaluations. Insights may be unstable.")

    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');
section[data-testid="stMain"] {
    font-family: 'Plus Jakarta Sans', sans-serif;
    background: #f0f2f7 !important;
    color: #1a2035;
}
section[data-testid="stMain"] .block-container { padding: 1.2rem 1.6rem !important; max-width: 100% !important; }
.card {
    background:#fff;
    border:1px solid #e4e8f0;
    border-radius:16px;
    padding:1.3rem 1.4rem 1rem;
    height:100%;
    min-height:240px;
    box-shadow:0 2px 12px rgba(30,50,110,0.05);
    display:flex;
    flex-direction:column;
}
.card-label { font-size:0.67rem; font-family:'JetBrains Mono',monospace; letter-spacing:0.1em; text-transform:uppercase; color:#94a3b8; margin-bottom:3px; }
.card-title { font-size:0.92rem; font-weight:700; color:#0f172a; letter-spacing:-0.2px; margin-bottom:0.2rem; }
.card-header { display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:0.4rem; }
.card-icon { width:36px; height:36px; border-radius:10px; display:flex; align-items:center; justify-content:center; font-size:16px; flex-shrink:0; }
.metric-value { font-size:2.6rem; font-weight:800; letter-spacing:-2px; color:#0f172a; line-height:1; margin:0.5rem 0 0.35rem; }
.metric-sub { font-size:0.7rem; color:#94a3b8; margin-top:7px; font-family:'JetBrains Mono',monospace; }
.stat-grid { display:grid; grid-template-columns:1fr 1fr; gap:9px; margin-top:6px; }
.card .stat-grid { margin-top:auto; }
.stat-item { background:#f8fafc; border:1px solid #e9ecf3; border-radius:10px; padding:10px 12px; }
.stat-val { font-size:1.3rem; font-weight:800; color:#0f172a; letter-spacing:-1px; }
.stat-lbl { font-size:0.64rem; color:#94a3b8; font-family:'JetBrains Mono',monospace; margin-top:2px; }
.prog-row { margin-bottom:11px; }
.prog-label-row { display:flex; justify-content:space-between; font-size:0.73rem; color:#475569; margin-bottom:5px; font-weight:500; }
.prog-bar-bg { background:#e9ecf3; border-radius:4px; height:7px; overflow:hidden; }
.prog-bar-fill { height:100%; border-radius:4px; }
</style>""", unsafe_allow_html=True)

    _sec_ratings: dict = {}
    _role_sec_rtg: dict = {}

    for ev in evaluations:
        _role = ev["evaluator_role"]
        for _sec in ev["sections"]:
            _st = _sec.get("title", "General")
            for _q in _sec.get("questions", []):
                if not isinstance(_q, dict) or _q.get("type") != "rating":
                    continue
                _qid = str(_q.get("id", ""))
                _rmax = float(_q.get("rating_max", 5))
                _ans = ev["answers"].get(_qid)
                if _ans is None or _ans == "":
                    continue
                if isinstance(_ans, dict):
                    _ans = _ans.get("rating")
                if _ans is not None:
                    try:
                        _v = float(_ans)
                        _sec_ratings.setdefault(_st, []).append((_v, _rmax))
                        _role_sec_rtg.setdefault(_role, {}).setdefault(_st, []).append((_v, _rmax))
                    except (ValueError, TypeError):
                        pass

    section_names: list = []
    section_avgs_5: list = []
    for _sn, _rts in _sec_ratings.items():
        if _rts:
            _a5 = sum(_v / _m * 5 for _v, _m in _rts) / len(_rts)
            section_names.append(_sn)
            section_avgs_5.append(round(_a5, 2))

    _all_vals = [_v / _m * 5 for _rts in _sec_ratings.values() for _v, _m in _rts]
    _overall5 = round(sum(_all_vals) / len(_all_vals), 2) if _all_vals else 0.0

    _role_avgs: dict = {}
    for _role, _sdct in _role_sec_rtg.items():
        _role_avgs[_role] = [
            round(sum(_v / _m * 5 for _v, _m in _sdct[_sn]) / len(_sdct[_sn]), 2)
            if _sdct.get(_sn) else 0.0
            for _sn in section_names
        ]

    _total_evals = len(evaluations)
    _unique_roles = len(set(ev["evaluator_role"] for ev in evaluations))
    _unique_forms = len(set(ev["form_name"] for ev in evaluations))
    _total_ans = sum(len(v) for v in _sec_ratings.values())

    _ROLE_COLS = ["#3b82f6", "#6366f1", "#0ea5e9", "#8b5cf6", "#06b6d4", "#f59e0b"]
    _COMP_COLS = ["#3b82f6", "#6366f1", "#0ea5e9", "#8b5cf6", "#06b6d4", "#f59e0b", "#10b981", "#f43f5e"]
    _FILL_RGBA = [
        "rgba(59,130,246,0.10)", "rgba(99,102,241,0.10)",
        "rgba(14,165,233,0.10)", "rgba(139,92,246,0.10)",
        "rgba(6,182,212,0.10)", "rgba(245,158,11,0.10)"
    ]
    _PLOT_CFG = dict(displayModeBar=False)
    _BASE_LAY = dict(
        paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
        font_color="#64748b", font_family="Plus Jakarta Sans",
        margin=dict(l=0, r=0, t=4, b=0),
    )

    def _sparkline_fig(_y):
        _fig = go.Figure()
        _fig.add_trace(go.Scatter(
            x=list(range(len(_y))), y=_y, mode="lines",
            line=dict(color="#3b82f6", width=2.5),
            fill="tozeroy", fillcolor="rgba(59,130,246,0.08)",
        ))
        _fig.update_layout(
            **{**_BASE_LAY, "paper_bgcolor": "rgba(0,0,0,0)", "plot_bgcolor": "rgba(0,0,0,0)"},
            height=75,
        )
        _fig.update_xaxes(visible=False)
        _fig.update_yaxes(visible=False)
        return _fig

    def _radar_fig():
        if not section_names:
            return None
        _snames = section_names[:]
        _pad_idx = 0
        while len(_snames) < 3:
            _pad_idx += 1
            _snames.append("\u200b" * _pad_idx)
        _cats = _snames + [_snames[0]]
        _fig = go.Figure()
        for _i, (_r, _avgs) in enumerate(_role_avgs.items()):
            _padded = (_avgs + [0.0, 0.0])[:len(_snames)]
            _fig.add_trace(go.Scatterpolar(
                r=_padded + [_padded[0]], theta=_cats,
                fill="toself", name=_r,
                line=dict(color=_ROLE_COLS[_i % len(_ROLE_COLS)], width=2),
                fillcolor=_FILL_RGBA[_i % len(_FILL_RGBA)],
            ))
        _fig.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            polar=dict(
                bgcolor="#ffffff",
                radialaxis=dict(visible=True, range=[0, 5], tickfont_size=8, gridcolor="#e9ecf3", linecolor="#e9ecf3"),
                angularaxis=dict(tickfont_size=9, linecolor="#e9ecf3", gridcolor="#e9ecf3"),
            ),
            legend=dict(font_size=9, orientation="h", x=0.5, xanchor="center", y=-0.08, bgcolor="rgba(0,0,0,0)"),
            margin=dict(l=30, r=30, t=10, b=22), height=230,
            font_color="#475569",
        )
        return _fig

    _cr1, _cr2, _cr3 = st.columns([1.05, 1.1, 0.85], gap="medium")

    with _cr1:
        st.markdown(f"""
<div class="card">
  <div class="card-header">
    <div>
      <div class="card-label">Summarized Evaluation</div>
      <div class="card-title">Average Point</div>
    </div>
    <div class="card-icon" style="background:#eff6ff">⭐</div>
  </div>
  <div class="metric-value">{_overall5:.2f}</div>
  <div class="metric-sub">Scale: 1.0 – 5.0 · {len(section_names)} competence</div>
</div>""", unsafe_allow_html=True)
        if section_avgs_5:
            st.plotly_chart(_sparkline_fig(section_avgs_5), use_container_width=True, config=_PLOT_CFG, key=f"{key_prefix}_sparkline")

    with _cr2:
        st.markdown("""
<div class="card" style="min-height:112px;padding-bottom:0.9rem;">
  <div class="card-header" style="margin-bottom:0;">
    <div>
      <div class="card-label">Competence map</div>
      <div class="card-title">Evaluator roles</div>
    </div>
    <div class="card-icon" style="background:#f0f9ff">🎯</div>
  </div>
</div>""", unsafe_allow_html=True)
        _rfig = _radar_fig()
        if _rfig:
            st.plotly_chart(_rfig, use_container_width=True, config=_PLOT_CFG, key=f"{key_prefix}_radar")
        else:
            st.caption("Nincs értékelési adat a radarhoz.")

    with _cr3:
        st.markdown(f"""
<div class="card">
  <div class="card-header">
    <div>
      <div class="card-label">Overview</div>
      <div class="card-title">Evaluation Status</div>
    </div>
    <div class="card-icon" style="background:#f0fdf4">📊</div>
  </div>
  <div class="stat-grid">
    <div class="stat-item"><div class="stat-val">{_total_evals}</div><div class="stat-lbl">Evaluations</div></div>
    <div class="stat-item"><div class="stat-val">{_unique_roles}</div><div class="stat-lbl">Roles</div></div>
    <div class="stat-item"><div class="stat-val">{_total_ans}</div><div class="stat-lbl">Answers</div></div>
    <div class="stat-item"><div class="stat-val">{_unique_forms}</div><div class="stat-lbl">Forms</div></div>
  </div>
</div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    _cr4, _cr5 = st.columns([1, 1], gap="medium")
    with _cr4:
        if section_names:
            _prog = ""
            for _cn, _sc, _cl in zip(section_names, section_avgs_5, (_COMP_COLS * (len(section_names) // len(_COMP_COLS) + 1))):
                _pct = int(_sc / 5 * 100)
                _prog += f"""
<div class="prog-row">
  <div class="prog-label-row">
    <span>{_cn}</span>
    <span style="color:{_cl};font-weight:700">{_sc:.2f}&nbsp;/&nbsp;5</span>
  </div>
  <div class="prog-bar-bg"><div class="prog-bar-fill" style="width:{_pct}%;background:{_cl}"></div></div>
</div>"""
            st.markdown(f"""
<div class="card">
  <div class="card-header">
    <div><div class="card-label">Detailed</div><div class="card-title">Competences</div></div>
    <div class="card-icon" style="background:#faf5ff">🧩</div>
  </div>
  <div style="margin-top:10px">{_prog}</div>
</div>""", unsafe_allow_html=True)
        else:
            st.info("No rating questions found in the evaluations.")

    with _cr5:
        if _role_avgs:
            _rnames = list(_role_avgs.keys())
            _rovrl = [
                round(sum(_a for _a in _avgs if _a > 0) / max(sum(1 for _a in _avgs if _a > 0), 1), 2)
                for _avgs in _role_avgs.values()
            ]
            _rb_fig = go.Figure(data=[go.Bar(
                x=_rnames, y=_rovrl,
                marker_color=(_ROLE_COLS * (len(_rnames) // len(_ROLE_COLS) + 1))[:len(_rnames)],
                text=[f"{_v:.2f}" for _v in _rovrl],
                textposition="auto",
            )])
            _rb_fig.update_layout(
                **{**_BASE_LAY, "margin": dict(l=10, r=10, t=4, b=30), "paper_bgcolor": "#ffffff", "plot_bgcolor": "#ffffff"},
                yaxis=dict(range=[0, 5], title="Score (1-5)", showgrid=True, gridcolor="rgba(0,0,0,0.04)", zeroline=False, showline=False, tickfont_size=9),
                xaxis=dict(showgrid=False, zeroline=False, showline=False, tickfont_size=9),
                height=160,
            )
            st.markdown("""
            <div class="card" style="min-height:auto;padding-bottom:0.6rem;">
              <div class="card-header" style="margin-bottom:0;">
                <div><div class="card-label">By Roles</div><div class="card-title">Average Point</div></div>
                <div class="card-icon" style="background:#eff6ff">👥</div>
              </div>
            </div>""", unsafe_allow_html=True)
            st.plotly_chart(_rb_fig, use_container_width=True, config=_PLOT_CFG, key=f"{key_prefix}_role_bar")
        else:
            st.info("No evaluator role data available.")


def _render_grouped_answers(evaluations: list, key_prefix: str = "answers"):
    if not evaluations:
        st.info("No completed evaluations found for this employee in this campaign.")
        return

    # Build distinct forms by ID (preserving appearance order)
    seen_form_ids: set = set()
    form_ids: list = []
    form_name_by_id: dict = {}
    for ev in evaluations:
        if ev["form_id"] not in seen_form_ids:
            seen_form_ids.add(ev["form_id"])
            form_ids.append(ev["form_id"])
            form_name_by_id[ev["form_id"]] = ev["form_name"]

    # Form filter – only rendered when there are multiple forms
    if len(form_ids) > 1:
        selected_form_ids = st.multiselect(
            "Filter by form",
            options=form_ids,
            default=form_ids,
            format_func=lambda fid: form_name_by_id.get(fid, str(fid)),
            placeholder="Select forms to display…",
            key=f"{key_prefix}_form_filter",
        )
    else:
        selected_form_ids = form_ids

    filtered_evals = [ev for ev in evaluations if ev["form_id"] in selected_form_ids]

    if not filtered_evals:
        st.info("No evaluations match the selected forms.")
        return

    # Group: form → evaluator_role → [evaluations]
    evals_by_form: dict = {}
    for ev in filtered_evals:
        evals_by_form.setdefault(ev["form_name"], {}).setdefault(
            ev["evaluator_role"], []
        ).append(ev)

    for form_name, roles_dict in evals_by_form.items():
        st.subheader(f"📋 {form_name}")

        # Use first evaluation's sections as question template
        first_eval = next(iter(next(iter(roles_dict.values()))))
        sections = first_eval["sections"]

        if not sections:
            st.caption("_(no questions in this form)_")
        else:
            for role_name, role_evals in roles_dict.items():
                st.markdown(f"#### 👤 {role_name}")

                for section in sections:
                    sec_title = section.get("title", "")
                    questions = section.get("questions", [])
                    if sec_title:
                        st.markdown(f"**{sec_title}**")

                    for q in questions:
                        if not isinstance(q, dict):
                            continue
                        q_id = str(q.get("id", ""))
                        q_text = q.get("text", "Question")

                        with st.container(border=True):
                            st.markdown(f"**{q_text}**")
                            any_answer = False
                            for ev in role_evals:
                                answer = ev["answers"].get(q_id)
                                if answer is not None and answer != "":
                                    _render_answer(answer, q)
                                    any_answer = True
                            if not any_answer:
                                st.caption("_(no answers)_")

        st.divider()


# =========================================================
# EMPLOYEE RESULT VIEW
# =========================================================
if st.session_state.cr_view == "overall":
    _render_cr_breadcrumb()
    st.title("Campaign Overall Results")
    st.caption(f"Campaign: {st.session_state.cr_selected_campaign_name}")
    st.divider()

    campaign_id = st.session_state.cr_selected_campaign_id

    with db.connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT
                e.id,
                eval_tor.name          AS evaluator_name,
                eval_tor.org_role_name AS evaluator_role,
                f.id                   AS form_id,
                f.name                 AS form_name,
                e.answers,
                e.finish_date,
                f.questions
            FROM evaluation e
            JOIN organisation_employees eval_tor ON e.evaluator_id = eval_tor.id
            JOIN form f ON e.form_id = f.id
            WHERE e.campaign_id = %s
              AND e.status      = 'completed'
            ORDER BY f.name, eval_tor.org_role_name, eval_tor.name
        """, (campaign_id,))

        evaluations = []
        for r in cur.fetchall():
            answers = r[5]
            if isinstance(answers, str):
                try:
                    answers = json.loads(answers)
                except (json.JSONDecodeError, ValueError):
                    answers = {}
            elif answers is None:
                answers = {}

            content = _normalize_questions(r[7])

            evaluations.append({
                "id":             r[0],
                "evaluator_name": r[1],
                "evaluator_role": r[2] or "Unknown",
                "form_id":        r[3],
                "form_name":      r[4],
                "answers":        answers,
                "finish_date":    r[6],
                "sections":       content["sections"],
            })
        cur.close()

    _render_summary_dashboard(evaluations, key_prefix="overall")

    st.divider()
    if st.button("← Back to Campaign Results", key="btn_back_campaign_results_overall"):
        _go_back_to_campaign_results()

elif st.session_state.cr_view == "employee":

    _render_cr_breadcrumb()

    st.title(f"Employee Evaluation Results - {st.session_state.cr_selected_employee_name}")
    st.caption(f"Campaign: {st.session_state.cr_selected_campaign_name}")
    st.divider()

    employee_id = st.session_state.cr_selected_employee_id
    campaign_id = st.session_state.cr_selected_campaign_id

    # Fetch all completed evaluations for this employee in this campaign
    with db.connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT
                e.id,
                eval_tor.name          AS evaluator_name,
                eval_tor.org_role_name AS evaluator_role,
                f.id                   AS form_id,
                f.name                 AS form_name,
                e.answers,
                e.finish_date,
                f.questions
            FROM evaluation e
            JOIN organisation_employees eval_tor ON e.evaluator_id = eval_tor.id
            JOIN form f ON e.form_id = f.id
            WHERE e.evaluatee_id = %s
              AND e.campaign_id  = %s
              AND e.status       = 'completed'
            ORDER BY f.name, eval_tor.org_role_name, eval_tor.name
        """, (employee_id, campaign_id))

        evaluations = []
        for r in cur.fetchall():
            answers = r[5]
            if isinstance(answers, str):
                try:
                    answers = json.loads(answers)
                except (json.JSONDecodeError, ValueError):
                    answers = {}
            elif answers is None:
                answers = {}

            content = _normalize_questions(r[7])

            evaluations.append({
                "id":             r[0],
                "evaluator_name": r[1],
                "evaluator_role": r[2] or "Unknown",
                "form_id":        r[3],
                "form_name":      r[4],
                "answers":        answers,
                "finish_date":    r[6],
                "sections":       content["sections"],
            })
        cur.close()

    tab_answers, tab_results = st.tabs(["📝 Answers", "📊 Results"])

    with tab_answers:
        _render_grouped_answers(evaluations, key_prefix="answers_all")

    with tab_results:
        selected_emp_name = (st.session_state.cr_selected_employee_name or "").strip().lower()
        self_evaluations = [
            ev for ev in evaluations
            if (ev.get("evaluator_name") or "").strip().lower() == selected_emp_name
        ]
        non_self_evaluations = [
            ev for ev in evaluations
            if (ev.get("evaluator_name") or "").strip().lower() != selected_emp_name
        ]

        role_names: list = []
        for ev in non_self_evaluations:
            role_name = ev.get("evaluator_role") or "Unknown"
            if role_name not in role_names:
                role_names.append(role_name)

        # Keep the classic structure explicitly visible: Summary / Self Evaluation / Role feedback tabs
        subpage_labels = ["Summary", "Self Evaluation"] + [f"{r} Role Feedback" for r in role_names]
        if not role_names:
            subpage_labels.append("Role Feedback")

        subpages = st.tabs(subpage_labels)

        with subpages[1]:
            _render_grouped_answers(self_evaluations, key_prefix="results_self")

        for idx, role_name in enumerate(role_names):
            with subpages[idx + 2]:
                role_evals = [ev for ev in non_self_evaluations if ev.get("evaluator_role") == role_name]
                _render_grouped_answers(role_evals, key_prefix=f"results_role_{idx}")

        if not role_names:
            with subpages[2]:
                st.info("No role-based feedback available yet for this employee.")

        with subpages[0]:
            if not evaluations:
                st.info("No completed evaluations found for this employee in this campaign.")
            else:
                _render_summary_dashboard(evaluations, key_prefix="emp")

                # ── AI Analysis ──────────────────────────────────────────────────
                st.divider()
                st.subheader("🤖 AI Analysis")

                if st.button(
                    "Generate AI Results",
                    key="btn_generate_ai_results_emp",
                    type="primary",
                    help="Run the AI pipeline for this employee's completed evaluations.",
                ):
                    # Build employee-scoped payload (same structure as get_campaign_qa_json)
                    emp_name = st.session_state.cr_selected_employee_name
                    form_questions_emp: dict = {}
                    for ev in evaluations:
                        fn      = ev["form_name"]
                        ev_role = ev["evaluator_role"]
                        if fn not in form_questions_emp:
                            form_questions_emp[fn] = {}
                            for section in ev["sections"]:
                                sec_title = section.get("title", "General")
                                for q in section.get("questions", []):
                                    if not isinstance(q, dict):
                                        continue
                                    q_id   = str(q.get("id", ""))
                                    q_type = q.get("type", "text")
                                    entry: dict = {
                                        "question":      q.get("text", ""),
                                        "question_type": q_type,
                                        "competence":    sec_title,
                                        "answers":       [],
                                    }
                                    if q_type == "multiple_choice":
                                        entry["options"] = q.get("options", [])
                                    elif q_type == "slider_labels":
                                        entry["options"] = q.get("slider_options", [])
                                    form_questions_emp[fn][q_id] = entry

                        for q_id, q_entry in form_questions_emp[fn].items():
                            answer_raw = ev["answers"].get(q_id)
                            if isinstance(answer_raw, dict):
                                answer_raw = (
                                    answer_raw.get("rating")
                                    or answer_raw.get("choice")
                                    or answer_raw.get("text")
                                )
                            if answer_raw is not None and answer_raw != "":
                                q_entry["answers"].append({
                                    "evaluatee_name": emp_name,
                                    "evaluator_role": ev_role,
                                    "answer":         answer_raw,
                                })

                    employee_payload = {
                        "campaign_id":   st.session_state.cr_selected_campaign_id,
                        "campaign_name": st.session_state.cr_selected_campaign_name,
                        "forms": {
                            fn: list(qs.values())
                            for fn, qs in form_questions_emp.items()
                        },
                    }

                    _script_path = os.path.abspath(
                        os.path.join(
                            os.path.dirname(__file__),
                            "..", "..", "..",
                            "services", "result_generation", "main.py",
                        )
                    )

                    with st.spinner("Running AI analysis…"):
                        proc = subprocess.run(
                            [sys.executable, _script_path, json.dumps(employee_payload)],
                            capture_output=True,
                            text=True,
                            timeout=120,
                        )

                    if proc.returncode == 0:
                        st.success("✅ Analysis finished.")
                        try:
                            output   = json.loads(proc.stdout)
                            analysis = output.get("results", {}).get(emp_name, {})
                            if not analysis:
                                st.json(output)
                            elif "raw_response" in analysis:
                                st.code(analysis["raw_response"])
                            else:
                                with st.expander("🔍 Raw LLM output (JSON)", expanded=False):
                                    st.json(analysis)

                                def _join(val) -> str:
                                    """Accept str or list[str] and return a plain string."""
                                    if isinstance(val, str):
                                        return val
                                    if isinstance(val, list):
                                        return ", ".join(str(v) for v in val)
                                    return str(val)

                                role_based = analysis.get("role_based_analysis", {})
                                if isinstance(role_based, dict) and role_based:
                                    st.markdown("**👥 Role-based Analysis**")
                                    for role_name, role_data in role_based.items():
                                        st.markdown(f"### {role_name}")

                                        strengths = role_data.get("strengths", []) if isinstance(role_data, dict) else []
                                        if strengths:
                                            st.markdown("**💪 Strengths**")
                                            for s in strengths:
                                                comps = _join(s.get("competence", ""))
                                                st.markdown(f"- **{comps}**")
                                                for ev_text in s.get("evidence", []):
                                                    st.caption(f"  • {ev_text}")

                                        areas = role_data.get("areas_for_improvement", []) if isinstance(role_data, dict) else []
                                        if areas:
                                            st.markdown("**📈 Areas for Improvement**")
                                            for a in areas:
                                                theme = _join(a.get("theme", ""))
                                                st.markdown(f"- **{theme}**")
                                                for ev_text in a.get("evidence", []):
                                                    st.caption(f"  • {ev_text}")
                                        st.divider()
                                else:
                                    # Backward compatibility with old schema
                                    strengths = analysis.get("strengths", [])
                                    if strengths:
                                        st.markdown("**💪 Strengths**")
                                        for s in strengths:
                                            comps = _join(s.get("competence", ""))
                                            st.markdown(f"- **{comps}**")
                                            for ev_text in s.get("evidence", []):
                                                st.caption(f"  • {ev_text}")
                                    areas = analysis.get("areas_for_improvement", [])
                                    if areas:
                                        st.markdown("**📈 Areas for Improvement**")
                                        for a in areas:
                                            theme = _join(a.get("theme", ""))
                                            st.markdown(f"- **{theme}**")
                                            for ev_text in a.get("evidence", []):
                                                st.caption(f"  • {ev_text}")

                                top_strengths = analysis.get("top_strengths", [])
                                if isinstance(top_strengths, list) and top_strengths:
                                    st.markdown("**🏆 Top 3 Strengths**")
                                    for item in top_strengths[:3]:
                                        st.markdown(f"- {item}")

                                top_dev = analysis.get("top_development_areas", [])
                                if isinstance(top_dev, list) and top_dev:
                                    st.markdown("**🛠️ Top 3 Development Areas**")
                                    for item in top_dev[:3]:
                                        st.markdown(f"- {item}")

                                summary = analysis.get("summary", "")
                                if summary:
                                    st.markdown("**📝 Summary**")
                                    st.write(summary)
                                conf_level  = analysis.get("confidence_level", "")
                                conf_reason = analysis.get("confidence_reason", "")
                                if conf_level:
                                    conf_color = {"high": "🟢", "medium": "🟡", "low": "🔴"}.get(conf_level, "⚪")
                                    st.caption(f"{conf_color} Confidence: **{conf_level}** — {conf_reason}")


                                # ── Export to Excel (fills eval_output_sample.xlsx template) ──
                                def _auto_fit_columns(ws, min_width=10, max_width=80, padding=2):
                                    for col_cells in ws.columns:
                                        max_length = 0
                                        col_index = col_cells[0].column
                                        col_letter = get_column_letter(col_index)

                                        for cell in col_cells:
                                            try:
                                                if cell.value is not None:
                                                    cell_length = len(str(cell.value))
                                                    if cell_length > max_length:
                                                        max_length = cell_length
                                            except Exception:
                                                pass

                                        adjusted_width = min(max(max_length + padding, min_width), max_width)
                                        ws.column_dimensions[col_letter].width = adjusted_width

                                # ── Export to Excel (fills eval_output_sample.xlsx template) ──
                                def _build_ai_excel(
                                    emp_name: str,
                                    emp_email: str,
                                    emp_role: str,
                                    camp_name: str,
                                    submitted_count: int,
                                    ana: dict,
                                ) -> bytes:
                                    # Locate the template relative to this file's location:
                                    # src/ui/pages/results/ → up 4 levels → datafiles/
                                    _tmpl = os.path.abspath(os.path.join(
                                        os.path.dirname(__file__),
                                        "..", "..", "..", "..",
                                        "datafiles", "eval_output_sample.xlsx",
                                    ))
                                    _wb = openpyxl.load_workbook(_tmpl)
                                    _ws = _wb["Summary"] if "Summary" in _wb.sheetnames else _wb.active

                                    def _eval_overall_5(_ev: dict) -> float | None:
                                        _vals = []
                                        for _sec in _ev.get("sections", []) or []:
                                            for _q in _sec.get("questions", []) or []:
                                                if not isinstance(_q, dict) or _q.get("type") != "rating":
                                                    continue
                                                _qid = str(_q.get("id", ""))
                                                _ans = (_ev.get("answers", {}) or {}).get(_qid)
                                                if isinstance(_ans, dict):
                                                    _ans = _ans.get("rating")
                                                if _ans is None or _ans == "":
                                                    continue
                                                try:
                                                    _v = float(_ans)
                                                    _m = float(_q.get("rating_max", 5))
                                                    if _m > 0:
                                                        _vals.append((_v / _m) * 5)
                                                except (ValueError, TypeError):
                                                    pass
                                        if not _vals:
                                            return None
                                        return sum(_vals) / len(_vals)

                                    # Per-evaluator overall averages (0-5 scale)
                                    _person_vals: dict[str, list[float]] = {}
                                    _person_role: dict[str, str] = {}
                                    for _idx, _ev in enumerate(evaluations):
                                        _ov = _eval_overall_5(_ev)
                                        if _ov is None:
                                            continue
                                        _name = (_ev.get("evaluator_name") or "").strip() or f"evaluator_{_idx}"
                                        _role = (_ev.get("evaluator_role") or "Unknown").strip() or "Unknown"
                                        _person_vals.setdefault(_name, []).append(_ov)
                                        _person_role.setdefault(_name, _role)

                                    _person_avg: dict[str, float] = {
                                        _n: (sum(_vs) / len(_vs)) for _n, _vs in _person_vals.items() if _vs
                                    }
                                    _manager_person_avgs = [
                                        _avg
                                        for _n, _avg in _person_avg.items()
                                        if "manager" in (_person_role.get(_n, "").lower())
                                    ]
                                    _non_manager_person_avgs = [
                                        _avg
                                        for _n, _avg in _person_avg.items()
                                        if "manager" not in (_person_role.get(_n, "").lower())
                                    ]

                                    _mgr_avg = (
                                        round(sum(_manager_person_avgs) / len(_manager_person_avgs), 2)
                                        if _manager_person_avgs else 0.0
                                    )
                                    _non_mgr_avg = (
                                        round(sum(_non_manager_person_avgs) / len(_non_manager_person_avgs), 2)
                                        if _non_manager_person_avgs else 0.0
                                    )
                                    _all_eval_avg = (
                                        round(sum(_person_avg.values()) / len(_person_avg), 2)
                                        if _person_avg else 0.0
                                    )

                                    def _set_value_safe(_sheet, _coord: str, _value) -> None:
                                        for _rng in _sheet.merged_cells.ranges:
                                            if _coord in _rng:
                                                _sheet.cell(row=_rng.min_row, column=_rng.min_col).value = _value
                                                return
                                        _sheet[_coord] = _value

                                    def _get_style_target_cell(_sheet, _coord: str):
                                        for _rng in _sheet.merged_cells.ranges:
                                            if _coord in _rng:
                                                return _sheet.cell(row=_rng.min_row, column=_rng.min_col)
                                        return _sheet[_coord]

                                    # ── Write to individual cells ─────────────────
                                    _set_value_safe(_ws, "B5", emp_name)
                                    _set_value_safe(_ws, "B6", emp_role if emp_role else "")
                                    _set_value_safe(_ws, "B7", emp_email if emp_email else "")
                                    _set_value_safe(_ws, "B11", date.today().strftime("%Y-%m-%d"))
                                    _set_value_safe(_ws, "B12", camp_name if camp_name else "")
                                    _set_value_safe(_ws, "B13", str(submitted_count))
                                    _set_value_safe(_ws, "E12", _mgr_avg)
                                    _set_value_safe(_ws, "H7", _non_mgr_avg)
                                    _set_value_safe(_ws, "H12", round(_non_mgr_avg - _mgr_avg, 2))
                                    _set_value_safe(_ws, "K7", _all_eval_avg)

                                    # Build competence averages locally for Excel export (scale: 0-5)
                                    _section_ratings: dict[str, list[tuple[float, float]]] = {}
                                    for _ev in evaluations:
                                        for _sec in _ev.get("sections", []) or []:
                                            _sec_title = (_sec.get("title") or "General").strip() or "General"
                                            for _q in _sec.get("questions", []) or []:
                                                if not isinstance(_q, dict) or _q.get("type") != "rating":
                                                    continue
                                                _qid = str(_q.get("id", ""))
                                                _ans = (_ev.get("answers", {}) or {}).get(_qid)
                                                if isinstance(_ans, dict):
                                                    _ans = _ans.get("rating")
                                                if _ans is None or _ans == "":
                                                    continue
                                                try:
                                                    _v = float(_ans)
                                                    _m = float(_q.get("rating_max", 5))
                                                    if _m > 0:
                                                        _section_ratings.setdefault(_sec_title, []).append((_v, _m))
                                                except (ValueError, TypeError):
                                                    pass

                                    _section_names = list(_section_ratings.keys())
                                    _section_avgs_5 = [
                                        round(sum(_v / _m * 5 for _v, _m in _rts) / len(_rts), 2)
                                        for _rts in _section_ratings.values() if _rts
                                    ]

                                    _competencies = [str(_c).strip() for _c in _section_names if str(_c).strip()]

                                    def _copy_row_style(_sheet, _src_row: int, _dst_row: int) -> None:
                                        _sheet.row_dimensions[_dst_row].height = _sheet.row_dimensions[_src_row].height
                                        for _col in range(1, _sheet.max_column + 1):
                                            _src_cell = _sheet.cell(row=_src_row, column=_col)
                                            _dst_cell = _sheet.cell(row=_dst_row, column=_col)
                                            if _src_cell.has_style:
                                                _dst_cell._style = copy(_src_cell._style)

                                    if _competencies:
                                        _start_row = 17
                                        _avg_by_comp = {
                                            str(_name).strip(): _avg
                                            for _name, _avg in zip(_section_names, _section_avgs_5)
                                        }

                                        # Header row directly above first competency.
                                        # We only fill the "Average" column for now.
                                        _header_row = _start_row - 1
                                        _avg_col = None
                                        for _col_idx in range(1, _ws.max_column + 1):
                                            _hv = _ws.cell(row=_header_row, column=_col_idx).value
                                            if isinstance(_hv, str) and "average" in _hv.strip().lower():
                                                _avg_col = _col_idx
                                                break
                                        if _avg_col is None:
                                            _avg_col = 2  # fallback: column right next to competency names

                                        _extra_rows = max(len(_competencies) - 1, 0)

                                        if _extra_rows > 0:
                                            # ── 1) Merged range-ek mentése és bontása a beszúrás előtt ──
                                            _saved_merges = []
                                            for _rng in list(_ws.merged_cells.ranges):
                                                if _rng.min_row >= _start_row:
                                                    _saved_merges.append((
                                                        _rng.min_col, _rng.min_row,
                                                        _rng.max_col, _rng.max_row,
                                                    ))
                                                    _ws.unmerge_cells(str(_rng))

                                            # ── 2) Sorok beszúrása ──
                                            _ws.insert_rows(_start_row + 1, amount=_extra_rows)

                                            # ── 3) Mentett merge-ek visszaállítása eltolva ──
                                            for _mc, _mr, _xc, _xr in _saved_merges:
                                                _new_min = _mr + _extra_rows if _mr > _start_row else _mr
                                                _new_max = _xr + _extra_rows if _xr > _start_row else _xr
                                                # Ha a merge a start_row-t is tartalmazza,
                                                # csak az alsó határt toljuk
                                                if _mr == _start_row:
                                                    _new_max = _xr + _extra_rows
                                                try:
                                                    _ws.merge_cells(
                                                        start_row=_new_min, start_column=_mc,
                                                        end_row=_new_max, end_column=_xc,
                                                    )
                                                except Exception:
                                                    pass

                                        # ── 4) Stílus másolása az összes beszúrt sorra ──
                                        for _idx in range(1, len(_competencies)):
                                            _row_idx = _start_row + _idx
                                            _copy_row_style(_ws, _start_row, _row_idx)

                                        # ── 5) Kompetencianevek beírása ──
                                        for _idx, _comp_name in enumerate(_competencies):
                                            _row_idx = _start_row + _idx
                                            _set_value_safe(_ws, f"A{_row_idx}", _comp_name)
                                            _a_cell = _get_style_target_cell(_ws, f"A{_row_idx}")
                                            _a_cell.alignment = Alignment(
                                                horizontal=_a_cell.alignment.horizontal if _a_cell.alignment else None,
                                                vertical=_a_cell.alignment.vertical if _a_cell.alignment else "center",
                                                text_rotation=_a_cell.alignment.text_rotation if _a_cell.alignment else 0,
                                                wrap_text=True,
                                                shrink_to_fit=False,
                                                indent=_a_cell.alignment.indent if _a_cell.alignment else 0,
                                            )

                                            # Competency average into the "Average" column.
                                            _avg_val = _avg_by_comp.get(_comp_name)
                                            if _avg_val is not None:
                                                _avg_coord = f"{get_column_letter(_avg_col)}{_row_idx}"
                                                _set_value_safe(_ws, _avg_coord, float(_avg_val))

                                        # ── 6) Felesleges üres sorok törlése a kompetenciák után ──
                                        _last_comp_row = _start_row + len(_competencies) - 1
                                        _scan_from = _last_comp_row + 1
                                        _empty_count = 0
                                        for _r in range(_scan_from, _scan_from + 20):
                                            _cell_val = _ws.cell(row=_r, column=1).value
                                            if _cell_val is None or str(_cell_val).strip() == "":
                                                _empty_count += 1
                                            else:
                                                break
                                        if _empty_count > 0:
                                            _ws.delete_rows(_scan_from, amount=_empty_count)

                                        # ── 7) Merge comment/notes blocks right below competencies ──
                                        _block_start_row = _start_row + len(_competencies)
                                        _block_end_row = _block_start_row + 3

                                        def _ranges_intersect(_a1: int, _a2: int, _a3: int, _a4: int,
                                                              _b1: int, _b2: int, _b3: int, _b4: int) -> bool:
                                            return not (_a3 < _b1 or _b3 < _a1 or _a4 < _b2 or _b4 < _a2)

                                        def _merge_block_safe(_min_col: int, _max_col: int) -> None:
                                            for _rng in list(_ws.merged_cells.ranges):
                                                if _ranges_intersect(
                                                    _min_col, _block_start_row, _max_col, _block_end_row,
                                                    _rng.min_col, _rng.min_row, _rng.max_col, _rng.max_row,
                                                ):
                                                    _ws.unmerge_cells(str(_rng))
                                            _ws.merge_cells(
                                                start_row=_block_start_row,
                                                start_column=_min_col,
                                                end_row=_block_end_row,
                                                end_column=_max_col,
                                            )

                                        # A..E block
                                        _merge_block_safe(1, 5)
                                        # G..K block
                                        _merge_block_safe(7, 11)

                                        # ── 8) Fill merged blocks with summarized strengths / development areas ──
                                        def _normalize_items(_raw_list, _key: str) -> list[str]:
                                            _out = []
                                            _seen = set()
                                            for _it in _raw_list or []:
                                                if not isinstance(_it, dict):
                                                    continue
                                                _v = _it.get(_key, "")
                                                if isinstance(_v, list):
                                                    _v = ", ".join(str(x).strip() for x in _v if str(x).strip())
                                                _txt = str(_v).strip()
                                                if _txt and _txt.lower() not in _seen:
                                                    _seen.add(_txt.lower())
                                                    _out.append(_txt)
                                            return _out

                                        # STRICT source policy for Excel summary cells:
                                        # only use Top Strengths and Top Development Areas.
                                        _top_strengths = ana.get("top_strengths", [])
                                        _top_dev = ana.get("top_development_areas", [])

                                        _strength_items = (
                                            [str(_x).strip() for _x in _top_strengths if str(_x).strip()]
                                            if isinstance(_top_strengths, list)
                                            else []
                                        )
                                        _dev_items = (
                                            [str(_x).strip() for _x in _top_dev if str(_x).strip()]
                                            if isinstance(_top_dev, list)
                                            else []
                                        )

                                        def _set_summary_cell(_coord: str, _title: str, _items: list[str]) -> None:
                                            _cell = _get_style_target_cell(_ws, _coord)
                                            _bullet_text = "\n".join(f"• {_x}" for _x in (_items or ["N/A"]))
                                            try:
                                                _rich = CellRichText()
                                                _rich.append(TextBlock(InlineFont(b=True), f"{_title}\n"))
                                                _rich.append(_bullet_text)
                                                _cell.value = _rich
                                            except Exception:
                                                _cell.value = f"{_title}\n{_bullet_text}"

                                            _cell.alignment = Alignment(
                                                horizontal="left",
                                                vertical="top",
                                                wrap_text=True,
                                                shrink_to_fit=False,
                                            )

                                        _set_summary_cell(f"A{_block_start_row}", "Key Strengths", _strength_items)
                                        _set_summary_cell(f"G{_block_start_row}", "Areas for Development", _dev_items)

                                    # Ensure text fields fit in cells
                                    _fit_alignment = Alignment(
                                        horizontal="left",
                                        vertical="center",
                                        wrap_text=True,
                                        shrink_to_fit=True,
                                    )
                                    for _cell_ref in ("B5", "B6", "B7", "B12"):
                                        _get_style_target_cell(_ws, _cell_ref).alignment = _fit_alignment

                                    _auto_fit_columns(_ws)
                                    #_ws["B2"]  = camp
                                    #_ws["A4"]  = ana.get("summary", "")
                                    #_ws["A5"]  = ana.get("confidence_level", "")
                                    #_ws["B5"]  = ana.get("confidence_reason", "")

                                    # Strengths – starting at row 8
                                    # _row = 8
                                    # for _s in ana.get("strengths", []):
                                    #     _comp = _s.get("competence", "")
                                    #     _comp = ", ".join(_comp) if isinstance(_comp, list) else str(_comp)
                                    #     for _ev in _s.get("evidence", []):
                                    #         _ws[f"A{_row}"] = _comp
                                    #         _ws[f"B{_row}"] = str(_ev)
                                    #         _row += 1
                                    #
                                    # # Areas for Improvement – starting at row 20
                                    # _row = 20
                                    # for _a in ana.get("areas_for_improvement", []):
                                    #     _theme = _a.get("theme", "")
                                    #     _theme = ", ".join(_theme) if isinstance(_theme, list) else str(_theme)
                                    #     for _ev in _a.get("evidence", []):
                                    #         _ws[f"A{_row}"] = _theme
                                    #         _ws[f"B{_row}"] = str(_ev)
                                    #         _row += 1
                                    #
                                    # _auto_fit_columns(_ws)

                                    # ── Detailed answers sheet ─────────────────────────────────
                                    # Remove old legacy sheet completely
                                    if "Részletes értékelések" in _wb.sheetnames:
                                        _wb.remove(_wb["Részletes értékelések"])

                                    _detail_sheet_name = "Detailed Answers"
                                    if _detail_sheet_name in _wb.sheetnames:
                                        _ws_detail = _wb[_detail_sheet_name]
                                    else:
                                        _ws_detail = _wb.create_sheet(_detail_sheet_name)

                                    # Fixed headers/order requested
                                    _ws_detail["A1"] = "Competence"
                                    _ws_detail["B1"] = "Question"
                                    _ws_detail["C1"] = "Question Type"
                                    _ws_detail["D1"] = "Evaluator Role"
                                    _ws_detail["E1"] = "Answer (text/non-rating)"
                                    _ws_detail["F1"] = "Score (rating)"

                                    # Clear previous detail rows
                                    _max_existing = max(_ws_detail.max_row, 2)
                                    for _row_idx in range(2, _max_existing + 1):
                                        for _col_idx in range(1, 7):
                                            _ws_detail.cell(row=_row_idx, column=_col_idx).value = None

                                    _row_out = 2
                                    for _ev in evaluations:
                                        _role = _ev.get("evaluator_role", "")
                                        _answers = _ev.get("answers", {}) or {}

                                        for _section in _ev.get("sections", []) or []:
                                            _category = _section.get("title", "General")
                                            for _q in _section.get("questions", []) or []:
                                                if not isinstance(_q, dict):
                                                    continue

                                                _q_id = str(_q.get("id", ""))
                                                _q_text = _q.get("text", "")
                                                _q_type = _q.get("type", "text")

                                                _ans = _answers.get(_q_id)
                                                if isinstance(_ans, dict):
                                                    _ans = _ans.get("rating") or _ans.get("choice") or _ans.get("text")

                                                _ws_detail.cell(row=_row_out, column=1).value = _category
                                                _ws_detail.cell(row=_row_out, column=2).value = _q_text
                                                _ws_detail.cell(row=_row_out, column=3).value = _q_type
                                                _ws_detail.cell(row=_row_out, column=4).value = _role

                                                # E: free text/non-rating answer
                                                # F: rating score
                                                if _q_type in ("rating", "slider_labels"):
                                                    _ws_detail.cell(row=_row_out, column=5).value = ""
                                                    _ws_detail.cell(row=_row_out, column=6).value = "" if _ans is None else str(_ans)
                                                else:
                                                    _ws_detail.cell(row=_row_out, column=5).value = "" if _ans is None else str(_ans)
                                                    _ws_detail.cell(row=_row_out, column=6).value = ""

                                                _row_out += 1

                                    _auto_fit_columns(_ws_detail)

                                    _buf = io.BytesIO()
                                    _wb.save(_buf)
                                    _buf.seek(0)
                                    return _buf.getvalue()

                                # Lekérdezzük az alkalmazott adatait (role, email) és a benyújtott értékelések számát
                                _emp_id = st.session_state.cr_selected_employee_id

                                with db.connection() as _conn:
                                    _cur = _conn.cursor()
                                    # Alkalmazott adatai
                                    _cur.execute(
                                        "SELECT name, email, org_role_name FROM organisation_employees WHERE id = %s",
                                        (_emp_id,)
                                    )
                                    _emp_row = _cur.fetchone()
                                    _emp_name_full = _emp_row[0] if _emp_row else emp_name
                                    _emp_email    = _emp_row[1] if _emp_row else ""
                                    _emp_role     = _emp_row[2] if _emp_row else ""

                                    # Benyújtott értékelések száma (beleértve az önértékelést is)
                                    _cur.execute(
                                        """
                                        SELECT COUNT(*) FROM evaluation
                                        WHERE evaluatee_id = %s
                                          AND campaign_id = %s
                                          AND status = 'completed'
                                        """,
                                        (_emp_id, campaign_id)
                                    )
                                    _count_row = _cur.fetchone()
                                    _submitted_count = _count_row[0] if _count_row else 0
                                    _cur.close()

                                _xlsx_bytes = _build_ai_excel(
                                    emp_name=_emp_name_full,
                                    emp_email=_emp_email,
                                    emp_role=_emp_role,
                                    camp_name=st.session_state.cr_selected_campaign_name,
                                    submitted_count=_submitted_count,
                                    ana=analysis,
                                )
                                st.download_button(
                                    label="⬇️ Export to Excel",
                                    data=_xlsx_bytes,
                                    file_name=f"ai_analysis_{emp_name.replace(' ', '_')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="btn_export_ai_excel",
                                )
                        except (json.JSONDecodeError, ValueError):
                            st.code(proc.stdout)
                    else:
                        st.error("❌ Pipeline returned an error.")
                        with st.expander("Error details"):
                            st.code(proc.stderr or proc.stdout)

    st.divider()
    if st.button("← Back to Campaign Results", key="btn_back_campaign_results_bottom"):
        _go_back_to_campaign_results()


# =========================================================
# CAMPAIGN RESULTS VIEW
# =========================================================
else:
    _render_cr_breadcrumb()
    st.title("Campaign Results")

    # Load campaigns
    with db.session() as session:
        campaigns = campaign_repo.list_campaigns()
        for c in campaigns:
            session.expunge(c)
    campaign_options = ["-- Select a campaign --"]
    campaign_dict = {}
    for c in campaigns:
        campaign_options.append(c.name)
        campaign_dict[c.name] = c.id

    # Restore previously selected campaign index
    current_idx = 0
    if (
        st.session_state.cr_selected_campaign_name
        and st.session_state.cr_selected_campaign_name in campaign_options
    ):
        current_idx = campaign_options.index(st.session_state.cr_selected_campaign_name)

    selected_campaign = st.selectbox("Select Campaign", campaign_options, index=current_idx)

    if selected_campaign != "-- Select a campaign --":
        campaign_id = campaign_dict[selected_campaign]
        st.session_state.cr_selected_campaign_id   = campaign_id
        st.session_state.cr_selected_campaign_name = selected_campaign

        with db.connection() as conn:
            participants_df = _participants_df_for_campaign(conn, campaign_id)

            cur = conn.cursor()
            cur.execute(
                """
                SELECT COUNT(*)
                FROM evaluation
                WHERE campaign_id = %s AND status = 'completed'
                """,
                (campaign_id,),
            )
            total_completed = int(cur.fetchone()[0] or 0)
            cur.close()

        participant_count = int(len(participants_df))
        participants_with_feedback = int((participants_df["completed_evaluations"] > 0).sum()) if participant_count else 0
        participation_rate = (participants_with_feedback / participant_count * 100) if participant_count else 0.0

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Participants", participant_count)
        k2.metric("Completed Evaluations", total_completed)
        k3.metric("Participants with Feedback", participants_with_feedback)
        k4.metric("Coverage", f"{participation_rate:.0f}%")

        if total_completed < 3:
            st.warning("Low data quality: too few completed evaluations for reliable campaign-level insight.")

        st.divider()






























































































































































































        with st.container(border=True):

















































            st.caption("Filters")
































































































































































































































































































































































            f1, f2 = st.columns([2.5, 1.5])
            with f1:
                search_name = st.text_input(
                    "Search participant",
                    key="cr_filter_search_name",
                    placeholder="Type name or email...",
                ).strip().lower()

            roles = sorted(participants_df["role"].fillna("No role").astype(str).unique().tolist()) if participant_count else []
            with f2:
                selected_roles = st.multiselect(
                    "Role",
                    options=roles,
                    default=[],
                    key="cr_filter_roles",
                    placeholder="Select role(s)...",
                )

            f3, f4 = st.columns([3.0, 1.0])
            with f3:
                max_completed = int(participants_df["completed_evaluations"].max()) if participant_count else 0
                min_completed = st.slider(
                    "Minimum completed evaluations",
                    min_value=0,
                    max_value=max_completed if max_completed > 0 else 0,
                    value=0,
                    key="cr_filter_min_completed",
                )
            with f4:
                st.write("")
                reset = st.button("Reset filters", key="cr_reset_filters", use_container_width=True)

            if reset:
                st.session_state.cr_filter_search_name = ""
                st.session_state.cr_filter_roles = []
                st.session_state.cr_filter_min_completed = 0
                st.rerun()

        filtered_df = participants_df.copy()
        if participant_count:
            if search_name:
                filtered_df = filtered_df[
                    filtered_df["name"].str.lower().str.contains(search_name, na=False)
                    | filtered_df["email"].fillna("").str.lower().str.contains(search_name, na=False)
                ]
            if selected_roles:
                filtered_df = filtered_df[filtered_df["role"].isin(selected_roles)]
            filtered_df = filtered_df[filtered_df["completed_evaluations"] >= int(min_completed)]

        chips = []
        if search_name:
            chips.append(f"search: {search_name}")
        if selected_roles:
            chips.append(f"roles: {', '.join(selected_roles)}")
        if int(min_completed) > 0:
            chips.append(f"min completed: {int(min_completed)}")
        if chips:
            st.caption("Active filters: " + " | ".join(chips))

        csv_df = filtered_df.copy()
        export_xlsx_bytes = None
        if not csv_df.empty:
            csv_df["campaign_name"] = selected_campaign
            csv_df["generated_at"] = datetime.utcnow().isoformat()
            csv_df["active_filters"] = "; ".join(chips) if chips else "none"

            export_df = csv_df.rename(
                columns={
                    "completed_evaluations": "completed evaluations",
                    "campaign_name": "campaign",
                    "generated_at": "generated at",
                    "active_filters": "active filters",
                }
            )

            export_buffer = io.BytesIO()
            with pd.ExcelWriter(export_buffer, engine="openpyxl") as writer:
                export_df.to_excel(writer, index=False, sheet_name="Participants")
            export_xlsx_bytes = export_buffer.getvalue()

        with st.container(border=True):
            st.caption("Export")
            st.download_button(
                ":material/download: Export participants (Excel)",
                data=export_xlsx_bytes if export_xlsx_bytes is not None else b"",
                file_name=f"campaign_participants_{campaign_id}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="cr_export_participants_xlsx",
                use_container_width=True,
                disabled=(export_xlsx_bytes is None),
                type="secondary",
            )

        st.divider()

        if not filtered_df.empty:
            preview_df = filtered_df[["name", "email", "role", "completed_evaluations"]].copy()
            preview_df = preview_df.sort_values(by=["completed_evaluations", "name"], ascending=[False, True])
            preview_df = preview_df.rename(
                columns={
                    "name": "Name",
                    "email": "Email",
                    "role": "Role",
                    "completed_evaluations": "Completed evaluations",
                }
            )
            st.dataframe(preview_df, use_container_width=True, hide_index=True)

            st.divider()

            with st.container(border=True):
                st.caption("Navigation")
                if st.button(
                    ":material/assessment: Overall Results",
                    key="btn_overall_summary",
                    use_container_width=True,
                    type="primary",
                ):
                    st.session_state.cr_view = "overall"
                    st.rerun()

            st.divider()

            employees = filtered_df.to_dict("records")
            for emp in employees:
                with st.container(border=True):
                    col_name, col_role, col_arrow = st.columns([3, 1, 1])
                    with col_name:
                        st.markdown(f"**{emp['name']}**")
                        st.caption(emp.get("email") or "")
                    with col_role:
                        st.caption(emp["role"] if emp["role"] else "No role")
                        st.caption(f"Completed: {int(emp.get('completed_evaluations', 0))}")
                    with col_arrow:
                        if st.button("→", key=f"btn_{emp['id']}"):
                            st.session_state.cr_view                  = "employee"
                            st.session_state.cr_selected_employee_id   = emp["id"]
                            st.session_state.cr_selected_employee_name = emp["name"]
                            st.rerun()
        else:
            st.info("No participants match the selected filters.")

    else:
        st.session_state.cr_selected_campaign_id   = None
        st.session_state.cr_selected_campaign_name = None
        st.info("Please select a campaign to view participants.")
